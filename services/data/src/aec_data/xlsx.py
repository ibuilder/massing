"""Tiny XLSX writer wrapper so every export shares formatting (guide §8 → XLSX/CSV)."""
from __future__ import annotations

from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def write_sheets(path: str, sheets: dict[str, tuple[Sequence[str], Sequence[Sequence[Any]]]]) -> str:
    """sheets = { sheet_name: (headers, rows) }. Returns the path written."""
    wb = Workbook()
    wb.remove(wb.active)
    for name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        ws.append(list(headers))
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append(list(row))
        # rough autofit
        for col_idx, header in enumerate(headers, start=1):
            width = max(len(str(header)), 10)
            for row in rows[:200]:
                if col_idx - 1 < len(row) and row[col_idx - 1] is not None:
                    width = max(width, min(len(str(row[col_idx - 1])), 60))
            ws.column_dimensions[get_column_letter(col_idx)].width = width + 2
        ws.freeze_panes = "A2"
    wb.save(path)
    return path
