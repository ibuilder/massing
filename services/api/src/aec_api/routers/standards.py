"""ISO 19650 / openBIM standards endpoints — CDE container discipline + requirements register."""
from __future__ import annotations

from fastapi import APIRouter, Body, Depends, File, HTTPException, Response, UploadFile
from sqlalchemy.orm import Session

from .. import bim_kpi, bsdd, cde, ids_authoring, mcp_tools, openbim, openbim_quality, standards_expert
from ..db import get_db
from ..models import Project
from ..rbac import current_user, require_role

router = APIRouter()


def _project(db: Session, pid: str) -> Project:
    p = db.get(Project, pid)
    if not p:
        raise HTTPException(404, "project not found")
    return p


@router.get("/projects/{pid}/bep")
def bep_generate(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """BEP-GEN: the **BIM Execution Plan** generated from the project's live configuration (ISO 19650) —
    standards + classification, information requirements (EIR/BEP/AIR + IDS), the RACI responsibility
    matrix, CDE container state, the model's exchange formats, and the model-quality acceptance gates.
    Always current — it reflects present state, not a stale point-in-time document. 404 if no project."""
    from .. import bep
    _project(db, pid)
    return bep.generate(db, pid)


@router.get("/projects/{pid}/cde/status")
def cde_status(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """CDE container rollup (ISO 19650): state distribution WIP/Shared/Published/Archived,
    suitability spread, and CDE-discipline metrics (revision control, approval-status coverage,
    metadata completeness)."""
    _project(db, pid)
    return cde.status(db, pid)


@router.get("/projects/{pid}/info-requirements/register")
def requirements_register(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """The information-requirements register (OIR/AIR/PIR/EIR/BEP/MIDP/TIDP) with issued/draft
    counts and core-document coverage (EIR, BEP, AIR)."""
    _project(db, pid)
    return cde.requirements(db, pid)


@router.get("/projects/{pid}/info-requirements/cascade")
def requirements_cascade(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """The ISO 19650 requirement flow-down — OIR → PIR/AIR → EIR → MIDP/TIDP linked by each record's
    `derives_from` — as a tiered tree plus cascade health (orphans that don't trace up to
    organizational intent; links pointing the wrong way)."""
    _project(db, pid)
    return cde.cascade(db, pid)


@router.get("/projects/{pid}/info-requirements/delivery-plan")
def requirements_delivery_plan(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """The MIDP/TIDP delivery plan — information requirements against their programme dates with
    overdue / due-soon status, a per-month roll-up, the next deliverable, and LOIN-specification
    coverage (EN 17412 Level of Information Need)."""
    _project(db, pid)
    return cde.delivery_plan(db, pid)


@router.get("/projects/{pid}/cde/exchange-acceptance")
def cde_exchange_acceptance(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """ISO 19650-6 information-exchange acceptance — each exchanged (non-WIP) container reviewed against
    completeness / suitability / authorization / traceability, with the non-conforming ones flagged."""
    _project(db, pid)
    return cde.exchange_acceptance(db, pid)


@router.get("/projects/{pid}/openbim/quality")
def openbim_quality_scan(pid: str, use_case: str | None = None, db: Session = Depends(get_db),
                         _: str = Depends(require_role("viewer"))):
    """openBIM quality of the loaded model: LOIN per element, IFC export health, bSDD alignment, and
    (when ?use_case= names an IDS use case) IDS rule-compliance scoring. Needs a loaded model."""
    _project(db, pid)
    from .properties import _INDEX, _ensure_loaded
    _ensure_loaded(pid)
    idx = _INDEX.get(pid)
    if not idx:
        raise HTTPException(404, "no properties index for project — load a model first")
    specs = ids_authoring.specs_for_use_case(use_case) if use_case else None
    out = openbim_quality.summary(idx, specs)
    out["use_case"] = use_case
    return out


@router.get("/projects/{pid}/lod/matrix")
def lod_matrix(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """The target LOD matrix (stage x discipline x element category -> LOD 100..500), or the RIBA/AIA
    stage defaults when the register carries none."""
    _project(db, pid)
    from .. import lod
    return lod.matrix(db, pid)


@router.get("/projects/{pid}/lod/assessment")
def lod_assessment(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """Achieved-LOD assessment of the loaded model (inferred from LOIN facet completeness) against the
    target matrix. Returns targets only when no model is loaded."""
    _project(db, pid)
    from .. import lod
    from .properties import _INDEX, _ensure_loaded
    try:
        _ensure_loaded(pid)
    except Exception:                     # noqa: BLE001 — no model loaded is a valid (targets-only) state
        pass
    return lod.assess(db, pid, _INDEX.get(pid))


@router.get("/projects/{pid}/naming/conventions")
def naming_conventions(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """The document/container filename + drawing sheet-ID naming conventions the validator enforces."""
    _project(db, pid)
    from .. import naming
    return naming.conventions()


@router.get("/projects/{pid}/naming/validate")
def naming_validate(pid: str, name: str, kind: str = "container",
                    db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """Validate a single name against the convention. kind = container | sheet."""
    _project(db, pid)
    from .. import naming
    return naming.validate(name, kind)


@router.get("/projects/{pid}/naming/audit")
def naming_audit(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """Audit the CDE containers + drawing register for naming-convention compliance."""
    _project(db, pid)
    from .. import naming
    return naming.audit(db, pid)


def _idx_for(pid: str):
    from .properties import _INDEX, _ensure_loaded
    try:
        _ensure_loaded(pid)
    except Exception:                     # noqa: BLE001 — no model is a valid (empty-result) state
        pass
    return _INDEX.get(pid)


@router.get("/projects/{pid}/model/query/views")
def model_query_views(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """The saved model-analytics views (count by discipline / class / storey / type)."""
    _project(db, pid)
    from .. import model_query
    return {"views": model_query.saved_views()}


@router.get("/projects/{pid}/model/query")
def model_query_run(pid: str, view: str | None = None, group_by: str = "ifc_class", agg: str = "count",
                    quantity: str | None = None, db: Session = Depends(get_db),
                    _: str = Depends(require_role("viewer"))):
    """Analytics query over the loaded model — a saved ?view=, or ad-hoc group_by / agg=sum&quantity=."""
    _project(db, pid)
    from .. import model_query
    idx = _idx_for(pid)
    return model_query.run_saved(idx, view) if view else model_query.query(idx, group_by, agg, quantity)


@router.get("/projects/{pid}/model/select")
def model_select(pid: str, q: str, limit: int = 5000, db: Session = Depends(get_db),
                 _: str = Depends(require_role("viewer"))):
    """QUERY-DSL — select elements with a selector string (`IfcWall & Pset_WallCommon.FireRating=2HR &
    storey=L3`) → matching GUIDs + the parsed predicates. One grammar scopes clash runs, view filters,
    schedules, bulk edits, and MCP tools. Bad query → 422."""
    from fastapi import HTTPException

    from .. import query_dsl
    _project(db, pid)
    try:
        return query_dsl.select(_idx_for(pid), q, limit=min(max(limit, 1), 20000))
    except query_dsl.QueryError as e:
        raise HTTPException(422, str(e))


@router.post("/projects/{pid}/answer/cited-query")
def answer_cited_query(pid: str, payload: dict = Body(...), db: Session = Depends(get_db),
                       _: str = Depends(require_role("viewer"))):
    """CITED-ANSWER — a deterministic model query that returns a **CitedAnswer**: every claim traces to the
    GUIDs it is derived from, with a coverage %, an uncited-claim guard, and source-conflict surfacing. The
    first producer of the provenance contract shared by the AI command bar / RFI-QA / KG answers. Body:
    `{query: <QUERY-DSL>, property?: <Pset.Prop>}`. Bad query → 422."""
    from .. import cited_answer, persona_answer, query_dsl
    _project(db, pid)
    q = str(payload.get("query") or "").strip()
    if not q:
        raise HTTPException(422, "query is required")
    try:
        out = cited_answer.cited_query(_idx_for(pid), q, prop=payload.get("property") or None,
                                       model_id=pid)
    except query_dsl.QueryError as e:
        raise HTTPException(422, str(e))
    if payload.get("persona"):
        out = persona_answer.shape(out, str(payload["persona"]))
    return out


@router.get("/projects/{pid}/model/fill-matrix")
def model_fill_matrix(pid: str, min_count: int = 1, db: Session = Depends(get_db),
                      _: str = Depends(require_role("viewer"))):
    """FILL-MATRIX — a category × property fill-rate pivot over the model: per IFC class, which pset field is
    systematically blank, with the **blank GUIDs** per property (feed them + a value to a bulk edit — the
    analytics → selection → bulk-write loop) + the worst partially-filled fields. Empty result without a
    loaded model."""
    from .. import fill_matrix as fm
    _project(db, pid)
    return fm.matrix(_idx_for(pid), min_count=max(1, min_count))


@router.post("/projects/{pid}/progress/rollup")
def progress_rollup(pid: str, payload: dict = Body(...), db: Session = Depends(get_db),
                    _: str = Depends(require_role("viewer"))):
    """PROGRESS-ROLLUP — percent-complete from as-built element presence: given the design element set and
    the installed GUIDs, roll up % complete by IFC class · discipline · level · overall, **by count and by
    value**. Body: `{installed_guids, elements?}`; when `elements` is omitted the design set is derived from
    the model's property index."""
    from .. import progress_rollup as pr
    _project(db, pid)
    elements = payload.get("elements")
    if not elements:
        idx = _idx_for(pid) or {}
        elements = [{"guid": g, "ifc_class": e.get("ifc_class"), "storey": e.get("storey")}
                    for g, e in idx.items()]
    return pr.rollup(elements, payload.get("installed_guids") or [])


@router.post("/projects/{pid}/progress/capture-diff")
def progress_capture_diff(pid: str, payload: dict = Body(...), db: Session = Depends(get_db),
                          _: str = Depends(require_role("viewer"))):
    """SCAN-4D — the diff between two capture timestamps: newly installed per class/level, elements that
    *disappeared* (re-scan/rework flag), the progress delta and a daily rate. Body: `{installed_t1,
    installed_t2, t1?, t2?, elements?}`; elements derive from the model's property index when omitted."""
    from .. import progress_rollup as pr
    _project(db, pid)
    elements = payload.get("elements")
    if not elements:
        idx = _idx_for(pid) or {}
        elements = [{"guid": g, "ifc_class": e.get("ifc_class"), "storey": e.get("storey")}
                    for g, e in idx.items()]
    return pr.capture_diff(elements, payload.get("installed_t1") or [], payload.get("installed_t2") or [],
                           payload.get("t1"), payload.get("t2"))


@router.get("/projects/{pid}/rules")
def rules_get(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """RULE-LIB — the project's user-authored parametric rule library (falls back to the starter set
    until one is saved). Each rule = a scope selector + a require selector + a severity."""
    from .. import rule_library
    _project(db, pid)
    stored = rule_library.load(pid)
    return {"rules": stored or rule_library.STARTER_RULES, "seeded": not stored}


@router.put("/projects/{pid}/rules")
def rules_put(pid: str, rules: list[dict] = Body(..., embed=True),
              db: Session = Depends(get_db), _: str = Depends(require_role("editor"))):
    """Replace the project's rule library. Every rule's selectors are validated (QUERY-DSL) before
    anything is written — a bad selector rejects the whole save with 422."""
    from .. import rule_library
    _project(db, pid)
    try:
        saved = rule_library.save(pid, rules)
    except rule_library.QueryError as e:
        raise HTTPException(422, str(e))
    return {"saved": len(saved), "rules": saved}


@router.get("/projects/{pid}/rules/space-pack")
def space_pack_get(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """RULE-PACK FOLD — the project's per-IfcSpace rule pack (dimensional thresholds · needs-daylight ·
    needs-wet-wall), stored beside the rule library and folded into /rules/run. Empty until saved."""
    from .. import rule_library
    _project(db, pid)
    return {"pack": rule_library.load_space_pack(pid)}


@router.put("/projects/{pid}/rules/space-pack")
def space_pack_put(pid: str, pack: dict = Body(..., embed=True),
                   db: Session = Depends(get_db), _: str = Depends(require_role("editor"))):
    """Replace the space rule pack. Validated (sections, severities, numeric ranges, type caps)
    before anything is written — a bad pack rejects the whole save with 422."""
    from .. import rule_library
    _project(db, pid)
    try:
        saved = rule_library.save_space_pack(pid, pack)
    except rule_library.QueryError as e:
        raise HTTPException(422, str(e))
    return {"pack": saved}


@router.get("/projects/{pid}/rules/run")
def rules_run(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """Check the loaded model against the rule library → per-rule pass/fail + offending GUIDs + a
    by-severity rollup. When a space rule pack is stored and the project has a source IFC, the
    geometric space checks (dimensional / daylight / wet-wall, via the adjacency engine) run too and
    fold into the SAME rollup as `space:*` rows — one rule spine for elements AND spaces."""
    from .. import rule_library
    _project(db, pid)
    stored = rule_library.load(pid) or rule_library.STARTER_RULES
    out = rule_library.run(_idx_for(pid), stored)
    pack = rule_library.load_space_pack(pid)
    if pack:
        try:
            from ..deps import open_source_ifc
            space_rows = rule_library.run_space_pack(open_source_ifc(db, pid), pack)
        except HTTPException:
            out["space_rules_note"] = "space pack stored but no source IFC is loaded — space checks skipped"
        else:
            out["rules"] = (out.get("rules") or []) + space_rows
            out["total_rules"] = len(out["rules"])
            for row in space_rows:
                if row["status"] == "fail":
                    out["failing_rules"] = out.get("failing_rules", 0) + 1
                    out["total_violations"] = out.get("total_violations", 0) + row["failed"]
                    bs = out.setdefault("by_severity", {})
                    bs[row["severity"]] = bs.get(row["severity"], 0) + 1
    return out


# VIEW-TEMPLATES (R18) — reusable layered view presets: class visibility + isolate scope + stacked
# color rules, deterministically resolvable.
@router.get("/projects/{pid}/view-templates")
def view_templates_get(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """The project's saved view templates (empty until saved)."""
    from .. import view_templates as vt
    _project(db, pid)
    return {"templates": vt.load(pid)}


@router.put("/projects/{pid}/view-templates")
def view_templates_put(pid: str, templates: list[dict] = Body(..., embed=True),
                       db: Session = Depends(get_db), _: str = Depends(require_role("editor"))):
    """Replace the project's view templates. Every template is validated (selectors parse, colors are
    #rrggbb, caps hold) before anything is written — a bad template rejects the whole save with 422."""
    from .. import view_templates as vt
    _project(db, pid)
    try:
        saved = vt.save(pid, templates)
    except vt.QueryError as e:
        raise HTTPException(422, str(e))
    return {"saved": len(saved), "templates": saved}


@router.get("/projects/{pid}/view-templates/{tid}/resolve")
def view_template_resolve(pid: str, tid: str, db: Session = Depends(get_db),
                          _: str = Depends(require_role("viewer"))):
    """Resolve one template against the loaded model → the deterministic visible GUID list + color map
    (later rules win) — the one answer the viewer AND the drawing generators consume. 404 unknown id."""
    from .. import view_templates as vt
    _project(db, pid)
    t = next((x for x in vt.load(pid) if x["id"] == tid), None)
    if t is None:
        raise HTTPException(404, f"view template {tid!r} not found")
    return vt.resolve(_idx_for(pid), t)


# SMART-VIEWS — user-authored saved view presets (name + QUERY-DSL selector + isolate/color/hide).
@router.get("/projects/{pid}/smart-views")
def smart_views_get(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """The project's saved smart views (property-driven view presets for the viewer)."""
    from .. import smart_views
    _project(db, pid)
    views = smart_views.load(pid)
    return {"views": views, "count": len(views)}


@router.put("/projects/{pid}/smart-views")
def smart_views_put(pid: str, views: list[dict] = Body(..., embed=True),
                    db: Session = Depends(get_db), _: str = Depends(require_role("editor"))):
    """Replace the project's smart views. Every view's selector is validated (QUERY-DSL) before
    anything is written — a bad selector rejects the whole save with 422."""
    from .. import smart_views
    _project(db, pid)
    try:
        saved = smart_views.save(pid, views)
    except smart_views.QueryError as e:
        raise HTTPException(422, str(e))
    return {"saved": len(saved), "views": saved}


@router.get("/projects/{pid}/smart-views/{vid}/run")
def smart_views_run(pid: str, vid: str, db: Session = Depends(get_db),
                    _: str = Depends(require_role("viewer"))):
    """Resolve a saved view's selector to the matching GUIDs (the viewer isolates / colours / hides
    them). 404 if the view id isn't saved."""
    from .. import smart_views
    _project(db, pid)
    view = next((v for v in smart_views.load(pid) if v["id"] == vid), None)
    if not view:
        raise HTTPException(404, f"no smart view '{vid}'")
    return smart_views.run(_idx_for(pid), view)


# RULE-LIB-2: the geometric checks a property selector can't express. Starter set when no checks
# are posted — same seeded-defaults pattern as the property library.
_GEO_DEFAULTS = [
    {"kind": "clearance", "name": "Door approach clearance", "scope": "IfcDoor",
     "distance_m": 0.9, "severity": "high"},
    {"kind": "clear_width", "name": "Accessible door clear width", "scope": "IfcDoor",
     "min_m": 0.815, "severity": "medium"},
    {"kind": "escape_distance", "name": "Space within reach of a door", "scope": "IfcSpace",
     "exits": "IfcDoor", "max_m": 60.0, "severity": "medium"},
]
_GEO_MAX_CHECKS = 20


@router.post("/projects/{pid}/rules/geometry/run")
def rules_geometry_run(pid: str, payload: dict = Body(default={}), db: Session = Depends(get_db),
                       _: str = Depends(require_role("editor"))):
    """Geometric/relational rule checks over the model's baked AABBs (the clash broad-phase path):
    `clearance` (approach space along the thin axis), `escape_distance` (straight-line to the
    nearest exit), `clear_width` (accessible opening width). Body `{checks: [{kind, scope, …}]}`
    with QUERY-DSL selector strings for scope/exits/obstructions; omit for the starter set.
    Editor role — this bakes geometry, like /clash."""
    from aec_data import geometric_rules  # type: ignore

    from .. import query_dsl
    from ..deps import source_ifc_path

    checks = payload.get("checks") or _GEO_DEFAULTS
    if not isinstance(checks, list) or len(checks) > _GEO_MAX_CHECKS:
        raise HTTPException(422, f"checks must be a list of at most {_GEO_MAX_CHECKS}")
    idx = _idx_for(pid)

    def _sel(q) -> set[str] | None:
        if q is None:
            return None
        try:
            return set(query_dsl.select(idx, str(q), limit=20000)["guids"])
        except query_dsl.QueryError as e:
            raise HTTPException(422, f"bad selector: {e}")

    resolved = []
    for c in checks:
        if c.get("kind") not in geometric_rules.KINDS:
            raise HTTPException(422, f"kind must be one of {geometric_rules.KINDS}")
        for p in ("distance_m", "max_m", "min_m"):
            if c.get(p) is not None:
                try:
                    ok = 0 < float(c[p]) <= 1000
                except (TypeError, ValueError):
                    ok = False
                if not ok:
                    raise HTTPException(422, f"{p} must be a number in (0, 1000]")
        resolved.append({**c, "scope": _sel(c.get("scope") or ""),
                         "exits": _sel(c.get("exits")) if c.get("exits") is not None else set(),
                         "obstructions": _sel(c.get("obstructions"))})
    boxes = geometric_rules.bake_boxes(source_ifc_path(db, pid))
    return geometric_rules.run(boxes, resolved)


@router.get("/projects/{pid}/model/roundtrip.csv")
def roundtrip_export(pid: str, props: str, db: Session = Depends(get_db),
                     _: str = Depends(require_role("viewer"))):
    """XLSX-ROUNDTRIP — the GUID-keyed property table for editing in Excel/Sheets: one row per
    element, columns = guid, ifc_class, name + the requested `Pset.Prop` columns (comma-separated).
    Re-import via POST /model/roundtrip/diff → the set_props_by_guid recipe."""
    import csv
    import io

    from .. import model_query
    _project(db, pid)
    fields = [p.strip() for p in (props or "").split(",") if p.strip()][:20]
    if not fields:
        raise HTTPException(422, "name at least one Pset.Prop column via ?props=")
    idx = _idx_for(pid) or {}

    def _cell(v) -> str:
        s = "" if v is None else str(v)
        # CSV formula-injection guard: Excel executes =+-@ leads; the diff parser strips one "'".
        return "'" + s if s[:1] in ("=", "+", "-", "@") else s

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["guid", "ifc_class", "name", *fields])
    for g, e in idx.items():
        w.writerow([g, e.get("ifc_class", ""), _cell(e.get("name")),
                    *[_cell(model_query._val(e, f.replace(".", "::", 1))) for f in fields]])
    return Response(buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": 'attachment; filename="properties.csv"'})


@router.post("/projects/{pid}/model/roundtrip/diff")
async def roundtrip_diff(pid: str, file: UploadFile = File(...), db: Session = Depends(get_db),
                         _: str = Depends(require_role("viewer"))):
    """XLSX-ROUNDTRIP — DRY-RUN diff of an edited CSV/XLSX against the live property index: which
    cells would change (`{guid, pset, prop, old, new, dtype}`), which GUIDs are unknown. Nothing is
    written — apply the returned `changes` via the `set_props_by_guid` edit recipe (which republishes).
    `dtype` is inferred from the OLD value's type so numeric properties don't flip to strings."""
    import csv
    import io

    from .. import model_query
    _project(db, pid)
    idx = _idx_for(pid) or {}
    raw = await file.read()
    name = (file.filename or "").lower()
    if name.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = [["" if c is None else str(c) for c in r] for r in ws.iter_rows(values_only=True)]
    else:
        rows = list(csv.reader(io.StringIO(raw.decode("utf-8-sig", "replace"))))
    if not rows or "guid" not in [h.strip().lower() for h in rows[0]]:
        raise HTTPException(422, "first row must be a header containing a 'guid' column")
    header = [h.strip() for h in rows[0]]
    gcol = [h.lower() for h in header].index("guid")
    prop_cols = [(i, h) for i, h in enumerate(header)
                 if "." in h and i != gcol and h.lower() not in ("ifc_class", "name")]
    changes: list[dict] = []
    unknown: list[str] = []
    unchanged = 0
    checked = 0
    for row in rows[1:5001]:                          # bounded: a sheet, not a bulk-import channel
        if gcol >= len(row) or not str(row[gcol]).strip():
            continue
        guid = str(row[gcol]).strip()
        e = idx.get(guid)
        checked += 1
        if e is None:
            unknown.append(guid)
            continue
        for i, h in prop_cols:
            new = str(row[i]).strip() if i < len(row) else ""
            if new.startswith("'"):                   # shed the export's formula-injection guard
                new = new[1:]
            if new == "":
                continue                              # blank cell = no edit intended
            old = model_query._val(e, h.replace(".", "::", 1))
            old_s = "" if old is None else str(old)
            if new == old_s:
                unchanged += 1
                continue
            pset, prop = h.split(".", 1)
            dtype = ("bool" if isinstance(old, bool) else "int" if isinstance(old, int)
                     else "float" if isinstance(old, float) else "str")
            changes.append({"guid": guid, "pset": pset, "prop": prop,
                            "old": old_s or None, "new": new, "dtype": dtype})
    return {"checked": checked, "changes": changes[:1000], "truncated": len(changes) > 1000,
            "unknown_guids": unknown[:100], "unchanged": unchanged}


@router.post("/projects/{pid}/ci/run")
def ci_run(pid: str, create_topics: bool = False, db: Session = Depends(get_db),
           actor: str = Depends(require_role("editor"))):
    """MODEL-CI — run the model check pack (rule library, completeness, clash, pinned IDS, quantity
    drift) → a pass/warn/fail report + badge, persisted as the project's latest CI result. With
    `create_topics=true`, each FAILING check becomes an open coordination Topic (BCF-model), so a CI
    failure round-trips to Solibri / ACC / BIMcollab like any other issue."""
    from .. import audit, model_ci
    from ..models import Topic
    _project(db, pid)
    rep = model_ci.run(db, pid, _idx_for(pid))
    created = 0
    if create_topics:
        for chk in rep["checks"]:
            if chk["status"] == "fail":
                db.add(Topic(project_id=pid, type="issue", status="open", author=actor,
                             title=f"Model CI: {chk['label']} failing",
                             description=f"{chk['summary']} (check `{chk['key']}`, run {rep['ran_at']})"))
                created += 1
        if created:
            audit.record(db, action="model_ci.create_topics", actor=actor, method="POST",
                         path=f"/projects/{pid}/ci/run", detail={"created": created})
        db.commit()
    return {**rep, "created_topics": created}


@router.get("/projects/{pid}/ci/latest")
def ci_latest(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """The project's last MODEL-CI report (the badge source)."""
    from .. import model_ci
    _project(db, pid)
    return model_ci.latest(pid)


@router.get("/projects/{pid}/model/export.csv")
def model_export_csv(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """Export the model element table as CSV (columnar, one row per element)."""
    _project(db, pid)
    from .. import model_query
    return Response(model_query.to_csv(_idx_for(pid)), media_type="text/csv",
                    headers={"Content-Disposition": f'attachment; filename="model-{pid}.csv"'})


@router.get("/projects/{pid}/model/export.jsonld")
def model_export_jsonld(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """Export the model elements as a JSON-LD graph (bSDD-style vocab, GlobalId as @id)."""
    _project(db, pid)
    from .. import model_query
    return model_query.to_jsonld(_idx_for(pid))


def _index_shape(pid: str) -> dict:
    """Adapt the in-memory columnar index ({guid: record}) to the element-index shape the IFC5 writer
    consumes (the same shape ifc5_reader produces). Empty but well-formed when no model is loaded."""
    idx = _idx_for(pid) or {}
    elements = []
    for rec in idx.values():
        elements.append({
            "guid": rec.get("guid"), "ifc_class": rec.get("ifc_class"), "name": rec.get("name"),
            "type_name": rec.get("type_name"), "storey": rec.get("storey"),
            "psets": rec.get("psets") or {}, "qtos": rec.get("qtos") or {},
        })
    classes = sorted({e["ifc_class"] for e in elements if e["ifc_class"]})
    storeys = sorted({e["storey"] for e in elements if e["storey"]})
    from ..model_index import get_meta  # per-project {schema, project, counts, facets}
    meta = get_meta(pid) or {}
    return {
        "schema": "IFC5", "project": meta.get("project") or {"guid": None, "name": None},
        "counts": {"elements": len(elements), "classes": len(classes), "storeys": len(storeys)},
        "facets": {"classes": classes, "storeys": storeys}, "elements": elements,
    }


@router.get("/projects/{pid}/model/export.ifcx")
def model_export_ifcx(pid: str, flavor: str = "ifcjson", db: Session = Depends(get_db),
                      _: str = Depends(require_role("viewer"))):
    """Export the model's element+property layer as IFC5 JSON. `flavor`: `ifcjson` (buildingSMART
    ifcJSON, default, full-fidelity round-trip) or `ifcx` (OpenUSD-style IFCX node list). Geometry is
    out of scope until web-ifc / Fragments add IFC5 upstream — this is the data write path."""
    _project(db, pid)
    from .. import licensing
    licensing.require_export("ifcx", "IFC5 / ifcJSON")   # openBIM data-out — Commercial+ when enforced
    import sys
    from pathlib import Path
    data_src = Path(__file__).resolve().parents[4] / "data" / "src"
    if str(data_src) not in sys.path:
        sys.path.insert(0, str(data_src))
    from aec_data import ifc5_writer  # type: ignore
    fl = "ifcx" if flavor == "ifcx" else "ifcjson"
    data = ifc5_writer.to_bytes(_index_shape(pid), fl, indent=2)
    return Response(data, media_type="application/json",
                    headers={"Content-Disposition": f'attachment; filename="model-{pid}.{fl}.ifcx"'})


@router.get("/projects/{pid}/model/export.parquet")
def model_export_parquet(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """Export the model element table as Apache Parquet (columnar analytics — DuckDB / pandas / Polars).

    Needs the optional `pyarrow` dependency; returns 503 with a clear message when it isn't installed."""
    _project(db, pid)
    from .. import model_query
    try:
        data = model_query.to_parquet(_idx_for(pid))
    except RuntimeError as exc:  # pyarrow not installed
        raise HTTPException(503, str(exc))
    return Response(data, media_type="application/vnd.apache.parquet",
                    headers={"Content-Disposition": f'attachment; filename="model-{pid}.parquet"'})


@router.get("/projects/{pid}/model/columnar/stats")
def model_columnar_stats(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """Interning/columnar efficiency stats for the loaded model — dedup ratio + estimated RAM saved by
    the BimOpenSchema-style string-interned columnar form vs the per-element JSON index."""
    _project(db, pid)
    from .. import bim_columns
    return bim_columns.interning_stats(_idx_for(pid))


@router.get("/projects/{pid}/model/columnar/aggregate")
def model_columnar_aggregate(pid: str, group_by: str = "ifc_class",
                             db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """Columnar count group-by over the element table via pyarrow compute (vectorised, no row loop)."""
    _project(db, pid)
    from .. import bim_columns
    try:
        return bim_columns.aggregate(_idx_for(pid), group_by)
    except RuntimeError as exc:      # pyarrow absent
        raise HTTPException(503, str(exc))
    except ValueError as exc:
        raise HTTPException(400, str(exc))


@router.get("/projects/{pid}/model/export/params.parquet")
def model_export_params_parquet(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """Export the model's property/quantity set as an EAV Parquet table (the analytics-friendly store —
    query in DuckDB/pandas). Needs pyarrow; 503 if absent."""
    _project(db, pid)
    from .. import bim_columns
    try:
        data = bim_columns.params_parquet(_idx_for(pid))
    except RuntimeError as exc:
        raise HTTPException(503, str(exc))
    return Response(data, media_type="application/vnd.apache.parquet",
                    headers={"Content-Disposition": f'attachment; filename="model-{pid}-params.parquet"'})


@router.get("/projects/{pid}/bim-kpi/scorecard")
def bim_kpi_scorecard(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """The 10-category BIM KPI scorecard, graded from the CDE, model quality and the issue / asset /
    closeout records (categories with no inputs show 'n/a')."""
    _project(db, pid)
    return bim_kpi.scorecard(db, pid)


@router.get("/projects/{pid}/handover/acceptance")
def handover_acceptance(pid: str, db: Session = Depends(get_db), _: str = Depends(require_role("viewer"))):
    """Handover data-drop acceptance gate — the owner's checklist against the AIR (requirements,
    asset tags, as-builts, O&M, accepted completion certificate)."""
    _project(db, pid)
    return bim_kpi.handover_acceptance(db, pid)


@router.get("/projects/{pid}/standards/check")
def standards_check(pid: str, standard: str = "iso19650", db: Session = Depends(get_db),
                    _: str = Depends(require_role("viewer"))):
    """Standards-compliance check (iso19650 | cobie | ids | uniclass) against the project's own data:
    findings with the clause each references, recommendations, and a 0–100 readiness score."""
    _project(db, pid)
    return standards_expert.check(db, pid, standard)


@router.get("/bsdd/search")
def bsdd_search(q: str, dictionary: str | None = None, limit: int = 20,
                _: str = Depends(current_user)):
    """Free-text search the buildingSMART Data Dictionary for classes matching `q`
    (optionally scoped to one ?dictionary= URI). Reference-data lookup, not
    project-scoped. A bSDD outage surfaces as 502, not 500."""
    try:
        return {"classes": bsdd.search_classes(q, dictionary_uri=dictionary, limit=limit)}
    except RuntimeError as exc:
        raise HTTPException(502, "bSDD unavailable") from exc


@router.get("/bsdd/class")
def bsdd_class(uri: str, _: str = Depends(current_user)):
    """Fetch one bSDD class (with its properties) by full `uri`. 404 when the class
    isn't found; 502 when bSDD is unreachable."""
    try:
        cls = bsdd.get_class(uri)
    except RuntimeError as exc:
        raise HTTPException(502, "bSDD unavailable") from exc
    if cls is None:
        raise HTTPException(404, "class not found")
    return cls


@router.get("/openbim/capabilities")
def openbim_capabilities(_: str = Depends(current_user)):
    """The openBIM standards + version matrix this platform speaks — for each standard (IFC, BCF, IDS,
    bSDD, COBie, ISO 19650 CDE), which versions we can read and write. Derived from the live engines
    (BCF versions, IFC schemas), so it never drifts from what's actually implemented; a consumer/agent
    can ask 'do you read BCF 3.0?' without guessing."""
    return openbim.capabilities()


@router.get("/mcp/tools")
def mcp_tool_catalog(_: str = Depends(current_user)):
    """The tool catalog the MCP server exposes to external AI agents (name, description, input
    schema). The stdio server (services/api/mcp_server.py) drives these against a project."""
    return {"tools": mcp_tools.catalog(), "server": "services/api/mcp_server.py",
            "note": "Run the stdio MCP server and point Claude Desktop / Cursor at it to drive the "
                    "project by natural language. The MCP SDK (pip install 'mcp[cli]') is optional."}

