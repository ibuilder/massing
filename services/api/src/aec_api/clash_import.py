"""Import a Solibri / Navisworks (or any tabular) clash / coordination report from XLSX into the
`coordination_issue` module — each data row becomes a coordination issue (which already round-trips
to BCF and drops a pin). GCs receive these reports constantly from the BIM coordinator; this turns a
spreadsheet into tracked, model-anchored issues without re-keying.

Tolerant by design: it sniffs the header row, maps a wide set of column aliases (Solibri's
Name/Description/Status/Component GUIDs; Navisworks' Clash Name/Status/Grid Location/Item 1/Item 2)
onto our fields, and skips blank rows. Pure parse function (testable without a DB) + a DB writer."""
from __future__ import annotations

import io
import re
from typing import Any

# our coordination_issue fields -> accepted column-name aliases (lowercased, punctuation-stripped)
_ALIASES: dict[str, tuple[str, ...]] = {
    "subject": ("name", "title", "clash name", "clashname", "issue", "issue name", "clash", "subject", "topic"),
    "description": ("description", "comment", "comments", "details", "detail", "notes", "remark", "remarks"),
    "discipline": ("discipline", "category", "ruleset", "rule", "rule name", "trade", "type", "clash type"),
    "location": ("location", "grid location", "grid", "level", "storey", "story", "zone", "space", "room", "floor"),
    "priority": ("severity", "priority", "status", "criticality", "importance"),
    "guids": ("guid", "guids", "component", "components", "ifc guid", "ifcguid", "element", "elements",
              "item 1", "item 2", "item1", "item2", "item 1 id", "item 2 id", "object id", "global id"),
}
_SEVERITY_MAP = {
    "critical": "Critical", "blocker": "Critical", "severe": "Critical",
    "high": "High", "major": "High", "error": "High", "fail": "High",
    "medium": "Medium", "normal": "Medium", "moderate": "Medium", "warning": "Medium", "warn": "Medium",
    "low": "Low", "minor": "Low", "info": "Low", "trivial": "Low",
}
_GUID_RE = re.compile(r"[0-9A-Za-z_$]{22}")          # IFC base64 GlobalId
_MAX_ROWS = 5000                                     # cap imported issues (DoS guard on a huge sheet)
_MAX_SCAN = 200_000                                  # hard cap on rows scanned before giving up
_MAX_XML_BYTES = 50 * 1024 * 1024                    # cap the Navisworks XML upload before DOM parse (50 MB)


def _norm(s: Any) -> str:
    return re.sub(r"[^a-z0-9 ]", "", str(s or "").strip().lower())


def _map_priority(v: Any) -> str | None:
    t = _norm(v)
    for kw, val in _SEVERITY_MAP.items():
        if kw in t:
            return val
    return None


def parse_clash_xlsx(data: bytes) -> dict[str, Any]:
    """Parse an XLSX clash/coordination export -> {rows: [{subject, description, ...}], columns, sheet}."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    grid = []
    for row in ws.iter_rows(values_only=True):       # read_only streams rows -> bounded memory
        grid.append(list(row))
        if len(grid) >= _MAX_SCAN:
            break
    wb.close()
    if not grid:
        return {"rows": [], "columns": {}, "sheet": ws.title, "header_row": None}

    # find the header row: the first row (within the first 15) where >=2 cells map to a known field.
    # match a column to the field whose alias matches best (exact > longest whole-word match), so
    # "Component GUID" -> guids (not "component"), "Rule Name" -> discipline (not subject's "name").
    def _best_field(cell: Any) -> str | None:
        n = _norm(cell)
        if not n:
            return None
        best, best_key = None, (0, 0)
        for field, aliases in _ALIASES.items():
            for a in aliases:
                exact = (n == a)
                hit = exact or re.search(r"\b" + re.escape(a) + r"\b", n) is not None
                if hit:
                    key = (1 if exact else 0, len(a))
                    if key > best_key:
                        best, best_key = field, key
        return best

    def col_map(row: list) -> dict[int, str]:
        out: dict[int, str] = {}
        for i, cell in enumerate(row):
            f = _best_field(cell)
            if f is not None:
                out[i] = f
        return out

    header_idx, mapping = None, {}
    for r in range(min(15, len(grid))):
        m = col_map(grid[r])
        if len(set(m.values())) >= 2:
            header_idx, mapping = r, m
            break
    if header_idx is None:
        return {"rows": [], "columns": {}, "sheet": ws.title, "header_row": None}

    rows: list[dict[str, Any]] = []
    truncated = False
    for raw in grid[header_idx + 1:]:
        if len(rows) >= _MAX_ROWS:
            truncated = True
            break
        rec: dict[str, Any] = {}
        guids: list[str] = []
        for i, field in mapping.items():
            val = raw[i] if i < len(raw) else None
            if val is None or str(val).strip() == "":
                continue
            if field == "guids":
                guids += _GUID_RE.findall(str(val))
            elif field == "priority":
                p = _map_priority(val)
                if p:
                    rec["priority"] = p
                # keep the raw status text in the description trail too
                rec.setdefault("_status_raw", str(val).strip())
            elif field in rec:                          # e.g. two component columns -> append
                rec[field] = f"{rec[field]} · {str(val).strip()}"[:2000]
            else:
                rec[field] = str(val).strip()
        if not rec.get("subject") and not rec.get("description") and not guids:
            continue                                    # blank / separator row
        if not rec.get("subject"):
            rec["subject"] = (rec.get("description") or "Coordination issue")[:120]
        status_raw = rec.pop("_status_raw", None)
        if status_raw and status_raw.lower() not in (rec.get("priority") or "").lower():
            rec["description"] = (rec.get("description", "") + (f"\nStatus: {status_raw}" if rec.get("description") else f"Status: {status_raw}")).strip()
        if guids:
            rec["_guids"] = sorted(set(guids))
        rows.append(rec)
    return {"rows": rows, "columns": {v: k for k, v in mapping.items()}, "sheet": ws.title,
            "header_row": header_idx, "truncated": truncated}


def parse_clash_xml(data: bytes) -> dict[str, Any]:
    """Parse a **native Navisworks clash-report XML** export into the same row shape as
    ``parse_clash_xlsx``. Namespace-agnostic (Navisworks uses the ``smart:`` namespace) + tolerant:
    each ``clashresult`` becomes a row (name → subject, its ``clashtest`` name → discipline, clash type/
    distance/status → description); GUIDs are harvested by scanning descendant text + attributes for the
    22-char IFC GlobalId pattern. Untrusted uploaded XML → parsed with defusedxml (XXE/entity-hardened)."""
    from defusedxml.common import DefusedXmlException
    from defusedxml.ElementTree import ParseError
    from defusedxml.ElementTree import fromstring as _safe_fromstring

    def _local(tag: Any) -> str:
        return str(tag).rsplit("}", 1)[-1].lower()

    if len(data) > _MAX_XML_BYTES:                     # bound memory: cap before building the full DOM
        return {"rows": [], "columns": {}, "sheet": "Navisworks XML", "truncated": True,
                "error": f"clash XML exceeds the {_MAX_XML_BYTES // (1024 * 1024)} MB limit"}
    try:
        root = _safe_fromstring(data)
    except (ParseError, DefusedXmlException, ValueError):
        return {"rows": [], "columns": {}, "sheet": "Navisworks XML", "truncated": False,
                "error": "unparseable or unsafe XML"}

    rows: list[dict] = []
    truncated = False
    for tc in root.iter():
        if _local(tc.tag) != "clashtest":
            continue
        test_name = (tc.get("name") or "").strip()
        for cr in tc.iter():
            if _local(cr.tag) != "clashresult":
                continue
            if len(rows) >= _MAX_ROWS:
                truncated = True
                break
            name = (cr.get("name") or "Clash").strip()
            status = (cr.get("status") or "").strip()
            dist = cr.get("distance")
            desc_bits, blob = [], []
            for e in cr.iter():
                lt = _local(e.tag)
                if lt == "description" and e.text:
                    desc_bits.append(e.text.strip())
                if e.text:
                    blob.append(e.text)
                blob.extend(str(v) for v in e.attrib.values())
            guids = sorted(set(_GUID_RE.findall(" ".join(blob))))
            desc = "; ".join(d for d in desc_bits if d)
            extra = " · ".join(x for x in (f"distance {dist}" if dist else "",
                                           f"status {status}" if status else "") if x)
            row: dict[str, Any] = {"subject": name[:200]}
            if test_name:
                row["discipline"] = test_name[:120]
            full = " · ".join(x for x in (desc, extra) if x)
            if full:
                row["description"] = full[:1000]
            pr = _map_priority(status)
            if pr:
                row["priority"] = pr
            if guids:
                row["_guids"] = guids[:20]
            rows.append(row)
    return {"rows": rows, "columns": {"navisworks_xml": "clashresult"}, "sheet": "Navisworks XML",
            "truncated": truncated}


def _write_issues(db, pid: str, parsed: dict[str, Any], actor: str) -> dict[str, Any]:
    """Create a `coordination_issue` per parsed row (GUIDs -> element_guids so it anchors on the model)."""
    from . import modules as me
    if "coordination_issue" not in me.TABLES:
        return {"error": "coordination_issue module not installed", "imported": 0}
    created = 0
    for r in parsed["rows"]:
        guids = r.get("_guids")                       # read, don't pop — never mutate the caller's rows
        body: dict[str, Any] = {"data": {k: v for k, v in r.items() if not k.startswith("_")}}
        if guids:
            body["element_guids"] = guids
        me.create_record(db, "coordination_issue", pid, body, actor, "GC")
        created += 1
    return {"imported": created, "detected_columns": list(parsed["columns"].keys()),
            "sheet": parsed["sheet"], "rows_parsed": len(parsed["rows"]),
            "truncated": parsed.get("truncated", False)}


def import_clash_xlsx(db, pid: str, data: bytes, actor: str) -> dict[str, Any]:
    """Parse an XLSX clash export + create a `coordination_issue` per row."""
    return _write_issues(db, pid, parse_clash_xlsx(data), actor)


def import_clash_xml(db, pid: str, data: bytes, actor: str) -> dict[str, Any]:
    """Parse a native Navisworks clash-report XML export + create a `coordination_issue` per clash."""
    return _write_issues(db, pid, parse_clash_xml(data), actor)
