"""Report Center — catalog + every report renders a valid PDF and a valid Excel workbook.
Run: PYTHONPATH=src ./.venv/Scripts/python.exe test_reports.py"""
import io
import os

os.environ["DATABASE_URL"] = "sqlite:///./test_reports.db"
os.environ["STORAGE_DIR"] = "./test_storage_reports"
os.environ.pop("AEC_RBAC", None)
for f in ("./test_reports.db",):
    if os.path.exists(f):
        os.remove(f)

import openpyxl                                              # noqa: E402
from fastapi.testclient import TestClient                   # noqa: E402
from aec_api.main import app                                # noqa: E402


def mk(c, pid, key, data):
    r = c.post(f"/projects/{pid}/modules/{key}", json={"data": data})
    assert r.status_code in (200, 201), f"{key}: {r.status_code} {r.text[:160]}"
    return r.json()["id"]


with TestClient(app) as c:
    pid = c.post("/projects", json={"name": "Report Tower"}).json()["id"]
    cc = mk(c, pid, "cost_code", {"code": "03-3000", "description": "Concrete", "division": "03"})
    mk(c, pid, "budget", {"cost_code": cc, "description": "Concrete", "revised": 2_000_000})
    com = mk(c, pid, "commitment", {"description": "Concrete sub", "cost_code": cc, "amount": 1_800_000})
    c.post(f"/projects/{pid}/modules/commitment/{com}/transition", json={"action": "execute"})
    mk(c, pid, "prime_contract", {"name": "GMP", "type": "GMP", "value": 5_000_000, "overhead_pct": 5, "fee_pct": 4})
    mk(c, pid, "subcontract", {"vendor": "ACME Concrete", "trade": "Concrete", "value": 1_800_000})
    mk(c, pid, "schedule_activity", {"name": "Foundations", "start": "2026-02-01", "finish": "2026-04-30",
                                     "budget": 2_000_000, "percent": 40})
    mk(c, pid, "cor", {"subject": "Added steel", "amount": 92_500, "reason": "Design Change"})
    mk(c, pid, "rfi", {"subject": "Beam clash", "question": "Please advise.", "discipline": "Structural", "cost_impact": "Yes"})
    mk(c, pid, "incident", {"subject": "Near miss", "date": "2026-06-14", "classification": "Near Miss", "severity": "Near Miss"})

    cat = c.get("/reports").json()["reports"]
    ids = [r["id"] for r in cat]
    assert {"executive", "cost", "evm", "change_orders", "rfi", "contracts"} <= set(ids), ids

    for rid in ids:
        p = c.get(f"/projects/{pid}/reports/{rid}.pdf")
        assert p.status_code == 200 and p.content[:4] == b"%PDF" and len(p.content) > 1200, f"{rid} pdf: {p.status_code} {len(p.content)}"
        x = c.get(f"/projects/{pid}/reports/{rid}.xlsx")
        assert x.status_code == 200, f"{rid} xlsx: {x.status_code} {x.text[:120]}"
        wb = openpyxl.load_workbook(io.BytesIO(x.content))      # must be a valid workbook
        assert wb.sheetnames, f"{rid}: no sheets"

    # spot-check content: the cost report's Excel carries the concrete category + a total
    wb = openpyxl.load_workbook(io.BytesIO(c.get(f"/projects/{pid}/reports/cost.xlsx").content))
    vals = [str(cell.value) for row in wb[wb.sheetnames[-1]].iter_rows() for cell in row]
    assert any("TOTAL" in v for v in vals), "cost xlsx should have a TOTAL row"
    assert c.get(f"/projects/{pid}/reports/nope.pdf").status_code == 404

    # risk digest endpoint: headline + prioritized risks + drivers (rule-based without an AI key)
    dg = c.get(f"/projects/{pid}/risk-digest").json()
    assert "headline" in dg and isinstance(dg.get("risks"), list), dg
    assert "schedule" in dg["drivers"] and "cost" in dg["drivers"], dg

    # the cost report carries a bar chart (budget vs committed vs actual vs EAC) into the PDF
    from aec_api import reports as _rep
    from aec_api import reports_render as _rr           # renderers now live in reports_render
    cost_rep = _rep.build(__import__("aec_api.db", fromlist=["SessionLocal"]).SessionLocal(), pid, "cost")
    assert cost_rep.charts and cost_rep.charts[0]["kind"] == "bar", cost_rep.charts
    assert _rr._chart_drawing(cost_rep.charts[0]) is not None    # renders without error

    # dispatch parity: every catalog report has a builder and vice-versa (build/catalog can't drift)
    _dispatch = set(_rep._BUILDERS) | set(_rep._LOGS)
    assert set(_rep.REPORTS) == _dispatch, set(_rep.REPORTS).symmetric_difference(_dispatch)

print(f"REPORTS OK - {len(ids)} reports each render a valid PDF + Excel workbook (incl. charts); cost bar; 404 on unknown")
