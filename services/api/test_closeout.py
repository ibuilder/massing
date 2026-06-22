"""Turnover/closeout deliverables: the closeout package ZIP + module-log PDF.
Run: PYTHONPATH=src ./.venv/Scripts/python.exe test_closeout.py"""
import io
import os
import zipfile

os.environ["DATABASE_URL"] = "sqlite:///./test_closeout.db"
os.environ["STORAGE_DIR"] = "./test_storage_closeout"
os.environ["IFC_DIR"] = "./test_ifc_closeout"
os.environ.pop("AEC_RBAC", None)
for f in ("./test_closeout.db",):
    if os.path.exists(f):
        os.remove(f)

from fastapi.testclient import TestClient  # noqa: E402
from aec_api.main import app  # noqa: E402

H = {"X-User": "gc"}

with TestClient(app) as c:
    pid = c.post("/projects", json={"name": "Turnover Tower"}, headers=H).json()["id"]
    # a generated model so the package has an as-built IFC + COBie/QTO/spaces
    g = c.post(f"/projects/{pid}/generate/massing",
               json={"lot_width": 30, "lot_depth": 20, "far": 2.0, "units": True}, headers=H)
    assert g.status_code == 200, g.text

    # closeout records (subject alias fills each module's title field)
    for mod, subj in [("commissioning", "HVAC Cx"), ("om_manual", "O&M"), ("warranty", "Roof 10yr"),
                      ("as_built", "Structural"), ("asset_register", "AHU-1"),
                      ("completion_certificate", "Substantial completion")]:
        r = c.post(f"/projects/{pid}/modules/{mod}", json={"data": {"subject": subj}}, headers=H)
        assert r.status_code == 201, (mod, r.text)
    # a couple RFIs for the log
    for s in ("Beam clash", "Door mismatch"):
        c.post(f"/projects/{pid}/modules/rfi", json={"data": {"subject": s, "question": "?"}}, headers=H)

    # --- closeout package ZIP -------------------------------------------------
    z = c.get(f"/projects/{pid}/closeout/package.zip", headers=H)
    assert z.status_code == 200 and z.headers["content-type"] == "application/zip", z.status_code
    zf = zipfile.ZipFile(io.BytesIO(z.content))
    names = zf.namelist()
    assert "as-built/model.ifc" in names, names
    assert "data/cobie.xlsx" in names and "data/qto.xlsx" in names and "data/spaces.xlsx" in names, names
    assert "report/status.pdf" in names, names
    assert "closeout/manifest.json" in names, names
    import json
    manifest = json.loads(zf.read("closeout/manifest.json"))
    assert manifest["closeout"]["warranty"][0]["title"] == "Roof 10yr", manifest["closeout"]["warranty"]
    assert len(manifest["closeout"]["completion_certificate"]) == 1

    # --- module log PDF -------------------------------------------------------
    log = c.get(f"/projects/{pid}/modules/rfi/log.pdf", headers=H)
    assert log.status_code == 200 and log.content[:4] == b"%PDF", log.status_code
    assert len(log.content) > 800, len(log.content)

    # --- COBie enrichment: closeout records fold into the workbook ------------
    import io as _io, zipfile as _zip
    import openpyxl  # ships with the xlsx writer
    cobie = c.get(f"/projects/{pid}/exports/cobie.xlsx", headers=H)
    assert cobie.status_code == 200, cobie.text
    wb = openpyxl.load_workbook(_io.BytesIO(cobie.content))
    assert "Warranty" in wb.sheetnames and "System" in wb.sheetnames, wb.sheetnames
    assert "Asset" in wb.sheetnames and "Document" in wb.sheetnames, wb.sheetnames

    # --- warranty expiry tracking ---------------------------------------------
    from datetime import date, timedelta
    soon = (date.today() + timedelta(days=30)).isoformat()
    past = (date.today() - timedelta(days=5)).isoformat()
    c.post(f"/projects/{pid}/modules/warranty", json={"data": {"name": "HVAC 1yr", "vendor": "Acme", "expires": soon}}, headers=H)
    c.post(f"/projects/{pid}/modules/warranty", json={"data": {"name": "Sealant", "vendor": "Bo", "expires": past}}, headers=H)
    w = c.get(f"/projects/{pid}/warranties/expiring?within_days=90", headers=H).json()
    assert any(x["name"] == "HVAC 1yr" for x in w["expiring"]), w["expiring"]
    assert any(x["name"] == "Sealant" for x in w["expired"]), w["expired"]

    # --- multi-period pay app + lien waiver (realistic flow: bill -> waiver -> advance) ---
    c.post(f"/projects/{pid}/modules/sov", json={"data": {"item_no": "01", "description": "Concrete",
           "scheduled_value": 1_000_000, "completed_this": 250_000, "retainage_pct": 5}}, headers=H)
    # period 1 lien waiver = current payment due (completed - retainage)
    lw = c.post(f"/projects/{pid}/cost/lien-waiver", json={"app_no": 1, "vendor": "Concrete Co"}, headers=H)
    assert lw.status_code == 201 and lw.json()["amount"] == 237_500, lw.text
    # advance the period: this -> prev, ready for application no. 2
    adv = c.post(f"/projects/{pid}/cost/pay-app/advance", headers=H).json()
    assert adv["advanced_lines"] == 1 and adv["next_application_no"] >= 1, adv
    g703 = c.get(f"/projects/{pid}/cost/g703", headers=H).json()
    assert g703["totals"]["prev"] == 250_000 and g703["totals"]["this"] == 0, g703["totals"]

print("CLOSEOUT OK - package.zip + log.pdf; COBie folds Warranty/System/Asset/Document tabs; "
      "warranty expiry tracking (expiring + expired); pay-app advance rolls this->prev; lien waiver")
